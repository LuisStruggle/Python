#!/usr/bin/env python3

"""
数据库
"""
import pickle
integers = [1, 2, 3, 4, 5]
f = open("dataxlh.dat", "wb")
pickle.dump(integers, f) # 对象系列化到文件
f.close()

integers = pickle.load(open("dataxlh.dat", "rb")) # 反序列化
print(integers)

import io #标准库中的一个模块，跟file功能类似，只不过是在内存中操作“文件”
class Book(object): #自定义一种类型
    def __init__(self,name):
        self.name = name
    def my_book(self):
        print("my book is: ", self.name)
pybook = Book("<from beginner to master>")
pybook.my_book()

file = io.BytesIO()
print(type(file))
pickle.dump(pybook, file, 1) # 1代表用二进制存储
print(file.getvalue()) #查看“文件”内容，注意下面不是乱码

file.seek(0) #找到对应类型
pybook2 = pickle.load(file)
pybook2.my_book()
file.close()

print("-----------------------------------------------------------------")
"""
pickle  模块已经表现出它足够好的一面了。不过，由于数据的复杂性， pickle  只能完成一
部分工作，在另外更复杂的情况下，它就稍显麻烦了。于是，又有了 shelve  。
"""
# shelve
import shelve
s=shelve.open("database.db")
s["name"]="liu"
s["sex"]="nan"
s["age"]="24"
s["contents"] = {"first":"base knowledge","second":"day day up"}
s.close()

# 上面是写入数据的过程，下面是读出数据的过程

s = shelve.open("database.db")
name = s["name"]
print(name)

print("-----------------------------------------------------------------")
# mysql
import pymysql
print(pymysql.__all__)
conn=pymysql.connect(host="localhost",user="root",passwd="123456",db="python",port=3306,charset="utf8")
cur = conn.cursor(cursor=pymysql.cursors.DictCursor) # 返回连接的游标对象。通过游标执行SQL查询并检查结果。游标比连接支持更多的方法，而且可能在程序中更好用。
# cur.execute("insert into student (name,age,sex) values (%s,%s,%s)",("liu","24","男"))
# conn.commit()

# 查询操作
cur.execute("select * from student")

lines = cur.fetchall()
for line in lines:
    print(line)

cur.close()
conn.close()

print("-----------------------------------------------------------------")
# SQLite的使用
import sqlite3
conn = sqlite3.connect("SQLite.db")
print(dir(conn))
cur = conn.cursor()
print(dir(cur))

# 创建数据库表
"""
create_table = "create table books (title text, author text, lang text)"
cur.execute(create_table)
"""
"""
cur.execute('insert into books values ("数据结构", "王三", "python")')
conn.commit()
"""
cur.close()
conn.close()

conn = sqlite3.connect("SQLite.db")
cur = conn.cursor()
cur.execute('select * from books')
print(cur.fetchall())

cur.close()
conn.close()

print("-----------------------------------------------------------------")
#openpyxl模块是解决Microsoft Excel 2007/2010之类版本中扩展名是Excel 2010 xlsx/xlsm/xltx/xltm的文件的读写的第三方库。
from openpyxl import Workbook
wb = Workbook()
ws = wb.active # 每个工作簿中，至少要有一个sheet，通过这条指令，就在当前工作簿中建立了一个sheet，并且它是当前正在使用的。

ws.title = "python" # ws  所引用的sheet对象名字就是"python"了。

print(wb.sheetnames)

b4 = ws['B4']
ws['B4'] = 4444
print(b4.value)

"""
要获得（或者建立并获得）某个cell对象，还可以使用下面方法：
>>> a1 = ws.cell("A1")
或者：
>>> a2 = ws.cell(row = 2, column = 1)
"""

cells = ws["A1":"C3"]
print(tuple(cells))

print(tuple(ws.rows))

wb.save("excel.xlsx")


from openpyxl import load_workbook
wb2 = load_workbook("excel.xlsx")
print(wb2.sheetnames)

ws_wb2 = wb2["python"]
for row in ws_wb2.rows:
    for cell in row:
        print(cell.value)



















